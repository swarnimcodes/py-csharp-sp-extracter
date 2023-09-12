import datetime
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side

"""
TODO: abstract off common functionality into a single function and
let the two functions for folder and file carry out small separate functions
"""

def tokenize_inl_query(inl_q: str) -> list[str]:
    temp_list: list[str] = inl_q.split()
    tbl_list: list[str] = [item for item in temp_list if (item.startswith("tbl"))]
    return tbl_list


def folder_analysis() -> None:
    folder: str = input("Enter folder path:\t")
    filelist: list[str] = []
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_fn: str = f"Folder_Analysis_{timestamp}.xlsx"

    sp_methods: list[str] = [
        "ExecuteNonQuerySP",
        "ExecuteNonQueryAsyncSP",
        "ExecuteReaderSP",
        "ExecuteReaderAsyncSP",
        "ExecuteScalarSP",
        "ExecuteScalarAsyncSP",
        "ExecuteDataSetSP",
    ]

    tbl_methods: list[str] = [
        "FillDropDownOnly",
    ]

    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws["A1"] = "File Path"
    ws['A1'].font = Font(bold=True)
    ws["B1"] = "SP Count"
    ws['B1'].font = Font(bold=True)
    ws['C1'] = "SP Line No."
    ws['C1'].font = Font(bold=True)
    ws["D1"] = "SP List"
    ws['D1'].font = Font(bold=True)
    ws["E1"] = "Table Count"
    ws['E1'].font = Font(bold=True)
    ws["F1"] = "Table List"
    ws['F1'].font = Font(bold=True)
    ws["G1"] = "Query Line No."
    ws['G1'].font = Font(bold=True)
    ws["H1"] = "Table Query"
    ws['H1'].font = Font(bold=True)

    # r=root, d=directories, f = files
    for r, d, f in os.walk(folder):
        for file in f:
            if ".cs" in file:
                filelist.append(os.path.join(r, file))

    for file in filelist:
        excel_row = []
        sp_count: int = 0
        table_count: int = 0
        sp_list: list[str] = []
        sp_ln: list[int] = []
        table_list: list[str] = []
        inl_ln: list[int] = []
        inl_query: list[str] = []
        print(file)
        with open(file, "r") as f:
            lines = f.readlines()
            line_num = 0
            for line in lines:
                line_num += 1
                if not line.startswith("//"):
                    if bool([ele for ele in sp_methods if (ele in line)]):
                        sp_list.extend(re.findall(r'"([^"]*)"', line))
                        sp_ln.append(line_num)
                        sp_count = sp_count + 1
                    elif bool([ele for ele in tbl_methods if (ele in line)]):
                        inl_query.append(re.findall(r'"([^"]*)"', line))
                        inl_ln.append(line_num)
                        match = re.search(r'"([^"]*)"', line)
                        if match:
                            m_tbl = match.group(1)
                            if not bool(re.search(r"\s", m_tbl)):
                                table_list.append(m_tbl)
                                table_count += 1
                            elif bool(re.search(r"\s", m_tbl)):
                                tbl_list = tokenize_inl_query(m_tbl)
                                table_list.extend(tbl_list)
                                table_count += 1
        print(
            f"Filename:\t{file}\nSP Count:\t{sp_count}\nSP List:\t{sp_list}\n"
            + f"Table Count:\t{table_count}\nTable List:\t{table_list}"
        )
        excel_row.append(file)
        excel_row.append(sp_count)
        excel_row.append("\n".join(map(str, sp_ln)))
        excel_row.append("\n".join(sp_list))
        excel_row.append(table_count)
        excel_row.append("\n".join(table_list))
        excel_row.append("\n".join(map(str, inl_ln)))
        excel_row.append("\n".join(map(str, inl_query)))
        ws.append(excel_row)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    border = Border(left=Side(style='thin'),
    right=Side(style='thin'),
    top = Side(style='thin'),
    bottom = Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    col_widths = {
        'A': 57,
        'B': 8,
        'C': 10,
        'D': 44,
        'E': 10,
        'F': 44,
        'G': 13,
        'H': 60,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    wb.save(excel_fn)
    print(f"Excel File generated at: {os.path.abspath(excel_fn)}")
    input("\nPress Enter to exit...\n")


def file_analysis() -> None:
    wb = Workbook()
    ws = wb.active
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_fn = f"File_Analysis_{timestamp}.xlsx"
    file = input("\n\nEnter file path or drag and drop file:\t")
    
    sp_methods: list[str] = [
        "ExecuteNonQuerySP",
        "ExecuteNonQueryAsyncSP",
        "ExecuteReaderSP",
        "ExecuteReaderAsyncSP",
        "ExecuteScalarSP",
        "ExecuteScalarAsyncSP",
        "ExecuteDataSetSP",
    ]

    tbl_methods: list[str] = [
        "FillDropDownOnly",
    ]

    # Headers
    ws["A1"] = "File Path"
    ws['A1'].font = Font(bold=True)
    ws["B1"] = "SP Count"
    ws['B1'].font = Font(bold=True)
    ws['C1'] = "SP Line No."
    ws['C1'].font = Font(bold=True)
    ws["D1"] = "SP List"
    ws['D1'].font = Font(bold=True)
    ws["E1"] = "Table Count"
    ws['E1'].font = Font(bold=True)
    ws["F1"] = "Table List"
    ws['F1'].font = Font(bold=True)
    ws["G1"] = "Query Line No."
    ws['G1'].font = Font(bold=True)
    ws["H1"] = "Table Query"
    ws['H1'].font = Font(bold=True)

    excel_row = []
    sp_count: int = 0
    table_count: int = 0
    sp_list: list[str] = []
    sp_ln: list[int] = []
    table_list: list[str] = []
    inl_ln: list[int] = []
    inl_query: list[str] = []

    with open(file, 'r') as f:
        lines = f.readlines()
        line_num = 0
        for line in lines:
            line_num = line_num + 1
            if not line.startswith("//"):
                if bool([ele for ele in sp_methods if (ele in line)]):
                    sp_list.extend(re.findall(r'"([^"]*)"', line))
                    sp_ln.append(line_num)
                    sp_count = sp_count + 1
                elif bool([ele for ele in tbl_methods if (ele in line)]):
                    inl_query.append(re.findall(r'"([^"]*)"', line))
                    inl_ln.append(line_num)
                    match = re.search(r'"([^"]*)"', line)
                    if match:
                        m_tbl = match.group(1)
                        if not bool(re.search(r"\s", m_tbl)):
                            table_list.append(m_tbl)
                            table_count += 1
                        elif bool(re.search(r"\s", m_tbl)):
                            tbl_list = tokenize_inl_query(m_tbl)
                            table_list.extend(tbl_list)
                            table_count += 1
    print(
        f"Filename:\t{file}\nSP Count:\t{sp_count}\nSP List:\t{sp_list}\n"
        + f"Table Count:\t{table_count}\nTable List:\t{table_list}"
    )
    excel_row.append(file)
    excel_row.append(sp_count)
    excel_row.append("\n".join(map(str, sp_ln)))
    excel_row.append("\n".join(sp_list))
    excel_row.append(table_count)
    excel_row.append("\n".join(table_list))
    excel_row.append("\n".join(map(str, inl_ln)))
    excel_row.append("\n".join(map(str, inl_query)))
    ws.append(excel_row)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
   
    border = Border(left=Side(style='thin'),
    right=Side(style='thin'),
    top = Side(style='thin'),
    bottom = Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    col_widths = {
        'A': 57,
        'B': 8,
        'C': 10,
        'D': 44,
        'E': 10,
        'F': 44,
        'G': 13,
        'H': 60,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
         
    wb.save(excel_fn)
    print(f"Excel File generated at: {os.path.abspath(excel_fn)}")
    input("\nPress Enter to exit...\n")
        


if __name__ == "__main__":
    print("Do you want to perform analysis on a folder or a file?\n")
    print("1. Folder\n2. File")
    choice = input("Enter your choice:\t")
    choice = int(choice)
    match choice:
        case 1:
            folder_analysis()
        case 2:
            file_analysis()
        case _:
            print("Error: Invalid choice.\n")
